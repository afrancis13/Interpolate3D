import matplotlib.pyplot as plt

from argparse import ArgumentParser
from mpl_toolkits.mplot3d import Axes3D
from openpyxl import load_workbook


class Interpolater(object):

    def __init__(self, txt, xlsx):
        '''
        Passes in a file in the forms of .xlsx and interpolates the
        percentage of occurrence of H and T values in a matrix.
        '''
        self.original_doc = txt
        self.interpolation_doc = xlsx
        self.map_h_t = {}
        self.H = []
        self.T = []
        self.percent = []

    def acquire_h_t_pairs(self):
        '''
        Acquires the H and T pairs from the user-inputted xlsx file.
        Fills the instance's H and T arrays.
        '''
        workbook = load_workbook(filename=self.interpolation_doc)
        valid_types = (float, int)
        for worksheet in workbook:
            for row in worksheet.rows:
                h_cell, t_cell = row[0], row[1]
                h, t = h_cell.value, t_cell.value
                if type(h) in valid_types and type(t) in valid_types:
                    self.H.append(h)
                    self.T.append(t)

    def acquire_matrix(self):
        '''
        Fills the map_h_t data structure with values of the user-inputted .txt
        file containing the H-T matrix. This is a dictionary that maps (H, T)
        to the percentage of occurrence (out of 100).
        '''
        with open(self.original_doc, 'r') as fp:
            initial = True
            for line in fp:
                if initial:
                    list_t = line.split('\t')
                    initial = False
                else:
                    line_split = line.split('\t')
                    h = float(line_split[0])
                    for i in range(1, len(line_split)):
                        try:
                            t = float(list_t[i])
                        except ValueError as e:
                            continue
                        try:
                            value = float(line_split[i])
                        except ValueError as e:
                            if line_split[i] == '':
                                self.map_h_t[(h, t)] = 0
                            continue
                        self.map_h_t[(h, t)] = value

    def acquire_closest_h_t(self, h, t):
        '''
        Acquires the closest upper and lower bounds on h and t in the matrix.
        These values will be used for the linear approximation in
        calculate_percentages.
        '''
        min_h_lower, diff_h_lower = None, -1 * float("inf")
        min_t_lower, diff_t_lower = None, -1 * float("inf")
        min_h_upper, diff_h_upper = None, float("inf")
        min_t_upper, diff_t_upper = None, float("inf")

        for (h_mat, t_mat) in self.map_h_t:
            diff_h = h_mat - h
            diff_t = t_mat - t
            if diff_h < 0 and diff_h > diff_h_lower:
                min_h_lower, diff_h_lower = h_mat, diff_h
            if diff_h >= 0 and diff_h < diff_h_upper:
                min_h_upper, diff_h_upper = h_mat, diff_h
            if diff_t < 0 and diff_t > diff_t_lower:
                min_t_lower, diff_t_lower = t_mat, diff_t
            if diff_t >= 0 and diff_t < diff_t_upper:
                min_t_upper, diff_t_upper = t_mat, diff_t

        return (min_h_lower, min_h_upper, min_t_lower, min_t_upper)

    def calculate_percentages(self):
    	'''
        Interpolates percentages from H and T by the following linear strategy:
        Find the H-T values that are "straddled by" the user-inputted H-T,
        and interpolate by taking a weighted average of the two H values and
        the two T values. Then, take a weighted average of the resulting
        averages. See the implementation.

        Methods to improve this approximation would be to perform polynomial
        regression on the values in the percentage matrix, but this methodology
        would likely result in a very small improvement.
        '''

        self.acquire_h_t_pairs()
        self.acquire_matrix()
        for i in range(len(self.H)):
            h_i, t_i = self.H[i], self.T[i]
            h_lower, h_upper, t_lower, t_upper = self.acquire_closest_h_t(h_i, t_i)

            if h_lower == None or h_upper == None or t_lower == None or t_upper == None:
                self.percent.append(0)
                continue

            percent_h_upper_t_upper = self.map_h_t[(h_upper, t_upper)]
            percent_h_upper_t_lower = self.map_h_t[(h_upper, t_lower)]
            percent_h_lower_t_upper = self.map_h_t[(h_lower, t_upper)]
            percent_h_lower_t_lower = self.map_h_t[(h_lower, t_lower)]

            slope_h_lower = float((percent_h_lower_t_upper - percent_h_upper_t_lower) / (t_upper - t_lower))
            slope_h_upper = float((percent_h_upper_t_upper - percent_h_upper_t_lower) / (t_upper - t_lower))
            weighted_average_h_lower = percent_h_lower_t_lower + slope_h_lower * (t_i - t_lower)
            weighted_average_h_upper = percent_h_upper_t_lower + slope_h_upper * (t_i - t_lower)

            slope_t = float((weighted_average_h_upper - weighted_average_h_lower) / (h_upper - h_lower))
            weighted_average_overall = weighted_average_h_lower + slope_t * (h_i - h_lower)

            self.percent.append(weighted_average_overall)

    def write_to_xlsx(self):
        '''
        Modifies the input Excel file. New saved file contains the calculated values!
        '''
        workbook = load_workbook(filename=self.interpolation_doc)
        for worksheet in workbook:
            for i in range(len(worksheet.rows)):
                cell = worksheet.cell(column=2, row=i)
                if i == 0:
                    cell.value = "Percent"
                else:
                    cell.value = self.percent[i - 1]

        workbook.save(filename=self.interpolation_doc)

    def clear(self):
        workbook = load_workbook(filename=self.interpolation_doc)
        for worksheet in workbook:
            for i in range(len(worksheet.rows)):
                cell = worksheet.cell(column=2, row=i)
                cell.value = ""

        workbook.save(filename=self.interpolation_doc)

    def plots(self):
        '''
        Plots a scatter of percentages with associated variable values.
        Only works for inputs of one or two variables, due to plotting
        constraints.

        Color coded:
            Blue: Inputted joint density data.
            Red: User requested joint density interpolation.
        '''
        fig_one = plt.figure()
        ax_one = Axes3D(fig_one)

        k, v = self.map_var_density.keys(), self.map_var_density.values()
        ordered_h = zip(*k)[0]
        ordered_t = zip(*k)[1]
        ordered_percent = v
        ax_one.scatter(ordered_h, ordered_t, ordered_percent, c='b')
        ax_one.scatter(self.H, self.T, self.percent, c='r')

        ax_one.set_xlabel('H')
        ax_one.set_ylabel('T')
        ax_one.set_zlabel('%')

        plt.show()


# Script

parser = ArgumentParser(description='Python joint density interpolation tool.')
parser.add_argument('density_matrix', help='Provided discrete matrix of joint densities.')
help_string = 'Values the user would like to interpolate based on the density matrix.'
parser.add_argument('values_to_interpolate', help=help_string)
parser.add_argument('-c', '--clear', action='store_true', help='Clear the percent column in all worksheets.')
parser.add_argument('-p', '--plot', action='store_true', help='Show plot of interpolated data with provided data.')

args = parser.parse_args()

original_doc = args.density_matrix
interpolation_doc = args.values_to_interpolate

i = Interpolater(original_doc, interpolation_doc)

if args.clear:
    i.clear()
    print "Finished clearing columns of the excel document..."
else:
    i.calculate_percentages()
    print "The percentages are as follows:\n" + str(i.percent)
    i.write_to_xlsx()
    if args.plot:
        i.plots()
